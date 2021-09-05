
import openpyxl
import requests

wb = openpyxl.Workbook()
from bs4 import BeautifulSoup
movie_list=[]
rank_list=[]
watched_list=[]
words_list=[]
director_list=[]
year_list=[]
country_list=[]
plot_list=[]
link_list=[]
a,b,c,d,e,f,g,h,ii,j=-1,-1,-1,-1,-1,-1,-1,-1,-1,-1
cnt=1

deep=1

def get_movies():
    global movie_list,director_list,country_list,plot_list,words_list,watched_list,link_list
    global a,b,c,d,e,f,g,h,ii,j,cnt
    headers={
    'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36',
    'Host' : 'movie.douban.com'
    }

    for i in range(0,deep):
        link='https://movie.douban.com/top250?start='+str(25*i) + '&filter='
        r=requests.get(link,headers=headers,timeout=10)
        print("正在爬取第",str(i+1),"页 状态码：",r.status_code)
        #print(r.text)
        soup=BeautifulSoup(r.text,'html.parser')

        ########################################## 电影名称
        div_list=soup.find_all('div',class_='hd')
        for each in div_list:
            movie=each.a.span.text.strip()
            movie_list.append(movie)
            print(movie)
            b=max(b,len(movie))          #记录电影名称长度最大值

        ########################################## 简语
        w = soup.find_all('div', class_='bd')
        for k in range(1,26):

            temp=str(w[k])
            pos=temp.find('''<span class="inq">''')
            # print(k," ",pos)
            if pos!=-1:
                pos2=temp.find("</span>",pos)
                words_list.append(temp[pos+18:pos2:])
                ii=max(ii,pos2-pos)
                #print(temp[pos+18:pos2:])
            else:
                words_list.append("无")
                #print("无")

        ########################################## 导演和演员、年份、国家、类别
        bd_list=soup.find_all('div',class_='bd')
        for i in bd_list:

            director=i.p.text.strip()
            Str=director

            pos=-1
            for i in range(0,len(Str)):
                if Str[i]=='1' or Str[i]=='2': #年份
                    pos=i
                    break
            director=director[0:pos:] #去除尾部
            year=Str[pos:pos+4:] #年份

            Str=Str[pos+6::]
            for i in range(0,len(Str)):
                if Str[i]=='/':
                    pos=i

            country=Str[:pos:] #国家
            plot=Str[pos+1::] #类别

            country_list.append(country)
            director_list.append(director)
            year_list.append(year)
            plot_list.append(plot)

            e = max(e, len(country)) #同理，记录最大长度
            c = max(c, len(director))
            d = max(d, len(year))
            f = max(f, len(plot))

        ########################################## 评分、已看人数
        star_list = soup.find_all('div', class_='star')
        for each in star_list:
            strs=str(each)
            if strs.find("average")!=-1:
                pos=strs.find("average")
                rank=strs[pos+9:pos+12:]
                rank_list.append(rank) #评分
                g = max(g, len(rank))
                #print(rank)

            if strs.find("<span>")!=-1:
                pos=strs.find("<span>")
                pos2=strs.find("人")
                watched=strs[pos+6:pos2:]
                watched_list.append(watched) #已看人数
                h = max(h, len(watched))
                #print(watched)

        ########################################## 链接
        links= soup.find_all('div', class_='pic')
        for each in links:
            link=each.a.attrs['href']
            link_list.append(link)
            j=max(j,len(link))




 ########################################## 写入表格
def sheet():
    global movie_list,director_list,year_list,country_list,plot_list,words_list,watched_list

    wb.create_sheet(index=0, title="douban")
    sheet=wb.active


    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = b*2+5 #根据实际情况调整
    sheet.column_dimensions['C'].width = c-5
    sheet.column_dimensions['D'].width = d+3
    sheet.column_dimensions['E'].width = e
    sheet.column_dimensions['F'].width = f*2+3
    sheet.column_dimensions['G'].width = g+5
    sheet.column_dimensions['H'].width = h+5
    sheet.column_dimensions['I'].width = ii*2-5
    sheet.column_dimensions['j'].width = j


    row0 = [u'排名', u'影名',u'主演',u'年份',u'国家',u'类别',u'评分',u'已看人数',u'简语',u'链接']
    for i in range(1,len(row0)+1):
        sheet.cell(1, i).value = row0[i-1] #表格第一行：排名、影名、主演。。。

    t=0
    ########################################## 按行写入
    for i in range(2,len(movie_list)+2):
        sheet.cell(i,1).value=str(i-1)
        sheet.cell(i,2).value=movie_list[i-2]
        sheet.cell(i, 7, rank_list[i - 2])
        sheet.cell(i, 8, watched_list[i - 2])
        sheet.cell(i, 9, words_list[i - 2])  #
        sheet.cell(i, 10, link_list[i - 2])

        if director_list[i+t-1]=="豆": #调整瑕疵
            t=t+1
        sheet.cell(i,3,director_list[i+t-1])
        sheet.cell(i,4,year_list[i+t-1])
        sheet.cell(i,5,country_list[i+t-1])
        sheet.cell(i,6,plot_list[i+t-1])



    wb.save('豆瓣排名.xls')
    print("********\n爬取完成！\n豆瓣排名.xls已保存在同目录文件夹下！")

def main():
    global deep
    print("爬取豆瓣电影 Top 250")
    print("请输入搜索深度（1～10）")
    deep=int(input())
    ans = deep*25
    print("将爬取",ans,"行信息，请稍后。。。")
    get_movies()
    sheet()



main()


